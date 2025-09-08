from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO

from core.permissions import IsSuperAdminOrSupportTeam
from .models import Group, GroupEmail, Tenant
from .serializers import (
    GroupSerializer, GroupEmailSerializer, GroupEmailModelSerializer, 
    BulkEmailUploadSerializer, ExcelUploadSerializer
)
from .admin_serializers import SuperAdminTenantSerializer


class IsSuperAdmin:
    """
    Custom permission class to allow only super admin users
    """
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role == 'super_admin'


class SuperAdminTenantViewSet(viewsets.ModelViewSet):
    """
    ViewSet for super admin to manage tenants
    """
    queryset = Tenant.objects.all()
    serializer_class = SuperAdminTenantSerializer
    permission_classes = [IsAuthenticated, IsSuperAdmin]
    filterset_fields = ['status', 'plan', 'is_active']
    search_fields = ['name', 'contact_email', 'contact_person', 'company_address']
    ordering_fields = ['name', 'created_at', 'status', 'plan']
    ordering = ['-created_at']

    def perform_create(self, serializer):
        if self.request.user.role != 'super_admin':
            raise PermissionDenied("Only super admin can create tenants")
        serializer.save()

from core.permissions import IsSuperAdminOrSupportTeam
from .models import Group, GroupEmail
from .serializers import (
    GroupSerializer, GroupEmailSerializer, GroupEmailModelSerializer, 
    BulkEmailUploadSerializer, ExcelUploadSerializer
)
class SuperAdminGroupViewSet(viewsets.ModelViewSet):
    """
    Super Admin Group ViewSet - Global access to all tenant groups
    """
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated, IsSuperAdminOrSupportTeam]
    filterset_fields = ['tenant', 'created_by', 'is_active']
    search_fields = ['name', 'description', 'tenant__name']
    ordering_fields = ['name', 'created_at', 'tenant__name']
    ordering = ['tenant__name', 'name']

    def get_queryset(self):
        """Super admin sees all groups across all tenants"""
        return Group.objects.all().select_related('tenant', 'created_by')

    def perform_create(self, serializer):
        """Allow super admin to create groups for any tenant"""
        # Super admin must specify tenant in the request data
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def bulk_add_emails(self, request, pk=None):
        group = self.get_object()
        serializer = BulkEmailUploadSerializer(data=request.data)
        if serializer.is_valid():
            contacts = serializer.validated_data['contacts']
            
            created_emails = []
            errors = []
            
            with transaction.atomic():
                for contact in contacts:
                    try:
                        email = contact['email']
                        name = contact.get('name', '')
                        organization = contact.get('organization', '')
                        
                        # Check if email already exists in this group
                        if GroupEmail.objects.filter(group=group, email=email).exists():
                            errors.append(f"Email {email} already exists in this group")
                            continue
                        
                        group_email = GroupEmail.objects.create(
                            group=group,
                            email=email,
                            name=name,
                            company=organization
                        )
                        created_emails.append(group_email)
                        
                        # Also create recipient in Recipients tab
                        from core.models_recipients import Recipient
                        recipient, recipient_created = Recipient.objects.get_or_create(
                            email=email,
                            tenant=group.tenant,
                            defaults={
                                'name': name,
                                'organization_name': organization,
                                'created_by': request.user
                            }
                        )
                        
                        # Add recipient to the contact group if it was created
                        if recipient_created:
                            from core.models_recipients import ContactGroup
                            contact_group, _ = ContactGroup.objects.get_or_create(
                                name=group.name,
                                tenant=group.tenant,
                                defaults={'created_by': request.user}
                            )
                            recipient.groups.add(contact_group)
                        
                    except Exception as e:
                        errors.append(f"Error adding {contact.get('email', 'unknown')}: {str(e)}")
            
            response_data = {
                'created_count': len(created_emails),
                'errors': errors,
                'message': f'Added {len(created_emails)} emails successfully'
            }
            
            return Response(response_data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def download_excel_template(self, request):
        """Download Excel template for email upload"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Email Template"
        
        # Headers
        headers = ['Name', 'Email', 'Organization Name', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Sample data
        sample_data = [
            ['John Doe', 'john@example.com', 'Example Corp', 'Sample contact'],
            ['Jane Smith', 'jane@example.com', 'Sample Inc', 'Another sample']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_name = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_name].width = adjusted_width
        
        # Save to bytes
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        response = HttpResponse(
            file_stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=group_email_template.xlsx'
        return response

    @action(detail=True, methods=['post'])
    def upload_excel(self, request, pk=None):
        group = self.get_object()
        serializer = ExcelUploadSerializer(data=request.data)
        
        if serializer.is_valid():
            contacts = serializer.validated_data['file']
            
            created_count = 0
            duplicate_count = 0
            
            with transaction.atomic():
                for contact in contacts:
                    email = contact['email']
                    name = contact.get('name', '')
                    organization = contact.get('organization', '')
                    
                    group_email, created = GroupEmail.objects.get_or_create(
                        group=group,
                        email=email,
                        defaults={
                            'name': name,
                            'company': organization
                        }
                    )
                    
                    if created:
                        created_count += 1
                        
                        # Also create recipient in Recipients tab
                        from core.models_recipients import Recipient
                        recipient, recipient_created = Recipient.objects.get_or_create(
                            email=email,
                            tenant=group.tenant,
                            defaults={
                                'name': name,
                                'organization_name': organization,
                                'created_by': request.user
                            }
                        )
                        
                        # Add recipient to the contact group if it was created
                        if recipient_created:
                            from core.models_recipients import ContactGroup
                            contact_group, _ = ContactGroup.objects.get_or_create(
                                name=group.name,
                                tenant=group.tenant,
                                defaults={'created_by': request.user}
                            )
                            recipient.groups.add(contact_group)
                    else:
                        duplicate_count += 1
            
            return Response({
                'message': f'Added {created_count} emails successfully from Excel file',
                'created': created_count,
                'duplicates': duplicate_count
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'])
    def bulk_remove_emails(self, request, pk=None):
        group = self.get_object()
        email_ids = request.data.get('email_ids', [])
        
        if not email_ids:
            return Response(
                {'error': 'email_ids is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        deleted_count = GroupEmail.objects.filter(
            group=group, 
            id__in=email_ids
        ).delete()[0]
        
        return Response(
            {'deleted_count': deleted_count}, 
            status=status.HTTP_200_OK
        )


class SuperAdminGroupEmailViewSet(viewsets.ModelViewSet):
    """
    Super Admin Group Email ViewSet - Global access to all group emails
    """
    serializer_class = GroupEmailModelSerializer
    permission_classes = [IsAuthenticated, IsSuperAdminOrSupportTeam]
    filterset_fields = ['group', 'group__tenant', 'is_active']
    search_fields = ['name', 'email', 'organization_name', 'group__name']
    ordering_fields = ['name', 'email', 'created_at', 'group__name']
    ordering = ['group__name', 'name']

    def get_queryset(self):
        """Super admin sees all group emails across all tenants"""
        return GroupEmail.objects.all().select_related('group', 'group__tenant', 'created_by')

    def perform_create(self, serializer):
        """Allow super admin to create group emails for any tenant"""
        group_email = serializer.save(created_by=self.request.user)
        
        # Also create recipient in Recipients tab
        from core.models_recipients import Recipient
        recipient, recipient_created = Recipient.objects.get_or_create(
            email=group_email.email,
            tenant=group_email.group.tenant,
            defaults={
                'name': group_email.name or '',
                'organization_name': group_email.company or '',
                'created_by': self.request.user
            }
        )
        
        # Add recipient to the contact group if it was created
        if recipient_created:
            from core.models_recipients import ContactGroup
            contact_group, _ = ContactGroup.objects.get_or_create(
                name=group_email.group.name,
                tenant=group_email.group.tenant,
                defaults={'created_by': self.request.user}
            )
            recipient.groups.add(contact_group)