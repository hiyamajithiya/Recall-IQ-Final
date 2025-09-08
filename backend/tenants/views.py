from rest_framework import generics, status, viewsets, mixins, parsers
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import pandas as pd
from datetime import datetime, timedelta # Import timedelta
from core.permissions import IsSuperAdmin, IsTenantAdmin, IsTenantOwner, IsTenantMember, IsSuperAdminOrSupportTeam
from .models import Tenant, TenantEmail, TenantMailSecret, Group, GroupEmail
from .serializers import (
    TenantSerializer, TenantEmailSerializer, TenantMailSecretSerializer,
    GroupSerializer, GroupEmailSerializer, GroupEmailModelSerializer, 
    BulkEmailUploadSerializer, ExcelUploadSerializer
)
from batches.models import Batch
from batches.serializers import BatchSerializer, BatchCreateSerializer, BatchActionSerializer
from .signals import manual_notify_trial_expiry, manual_notify_subscription_renewal
from core.tenant_notifications import send_tenant_notification


class TenantViewSet(viewsets.ModelViewSet):
    serializer_class = TenantSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.role in ['super_admin', 'support_team']:
            return Tenant.objects.all()
        elif self.request.user.tenant:
            return Tenant.objects.filter(id=self.request.user.tenant.id)
        return Tenant.objects.none()
    
    def get_permissions(self):
        if self.action == 'create':
            return [IsSuperAdminOrSupportTeam()]
        elif self.action == 'destroy':
            return [IsSuperAdmin()]
        elif self.action in ['update', 'partial_update']:
            # Super admin and support team can update any tenant
            if hasattr(self.request.user, 'role') and self.request.user.role in ['super_admin', 'support_team']:
                return [IsSuperAdminOrSupportTeam()]
            # Tenant users can only update their own tenant
            return [IsTenantAdmin(), IsTenantOwner()]
        elif self.action in ['send_trial_expiry_notification', 'send_renewal_notification', 'send_custom_notification']:
            return [IsSuperAdmin()]
        return [IsTenantMember()]

    # --- MODIFIED ---
    # This function is called when a new tenant is created via the API.
    def perform_create(self, serializer):
        """Override to automatically set a 30-day trial period on creation."""
        # Calculate the trial end date: today + 30 days
        trial_end = timezone.now() + timedelta(days=30)
        
        # Save the new tenant with the calculated trial_end_date
        # and set the initial status to 'trial'.
        serializer.save(
            trial_end_date=trial_end,
            status='trial'
        )
    # ----------------

    def perform_update(self, serializer):
        """Override to track who made the changes for notifications"""
        import logging
        logger = logging.getLogger(__name__)
        
        instance = serializer.instance
        instance._changed_by = self.request.user
        
        logger.info(f"🔧 perform_update called for tenant {instance.name} by {self.request.user.username}")
        logger.info(f"🔧 Setting _changed_by to: {instance._changed_by}")
        
        result = serializer.save()
        
        logger.info(f"🔧 serializer.save() completed")
        return result
    
    @action(detail=True, methods=['post'])
    def send_trial_expiry_notification(self, request, pk=None):
        """Manually send trial expiry notification to tenant"""
        try:
            tenant = self.get_object()
            result = manual_notify_trial_expiry(tenant.id, request.user)
            
            if result:
                return Response({
                    'message': f'Trial expiry notification sent successfully to {tenant.name}',
                    'tenant_id': tenant.id,
                    'tenant_name': tenant.name
                })
            else:
                return Response({
                    'error': 'Failed to send trial expiry notification'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'error': f'Error sending notification: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def send_renewal_notification(self, request, pk=None):
        """Manually send subscription renewal notification to tenant"""
        try:
            tenant = self.get_object()
            result = manual_notify_subscription_renewal(tenant.id, request.user)
            
            if result:
                return Response({
                    'message': f'Subscription renewal notification sent successfully to {tenant.name}',
                    'tenant_id': tenant.id,
                    'tenant_name': tenant.name
                })
            else:
                return Response({
                    'error': 'Failed to send renewal notification'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'error': f'Error sending notification: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def send_custom_notification(self, request, pk=None):
        """Send custom notification to tenant"""
        try:
            tenant = self.get_object()
            notification_type = request.data.get('type', 'generic_update')
            changes = request.data.get('changes', {})
            
            result = send_tenant_notification(tenant, notification_type, changes, request.user)
            
            if result:
                return Response({
                    'message': f'Custom notification sent successfully to {tenant.name}',
                    'tenant_id': tenant.id,
                    'tenant_name': tenant.name,
                    'notification_type': notification_type
                })
            else:
                return Response({
                    'error': 'Failed to send custom notification'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'error': f'Error sending notification: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ... The rest of your views.py file remains the same ...
# (TenantEmailViewSet, TenantMailSecretViewSet, Group ViewSets, etc.)

class TenantEmailViewSet(viewsets.ModelViewSet):
    serializer_class = TenantEmailSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]
    
    def get_queryset(self):
        if self.request.user.role == 'super_admin':
            return TenantEmail.objects.all()
        elif self.request.user.tenant:
            return TenantEmail.objects.filter(tenant=self.request.user.tenant)
        return TenantEmail.objects.none()
    
    def perform_create(self, serializer):
        if self.request.user.role != 'super_admin':
            serializer.save(tenant=self.request.user.tenant)
        else:
            serializer.save()


class TenantMailSecretViewSet(viewsets.ModelViewSet):
    serializer_class = TenantMailSecretSerializer
    permission_classes = [IsAuthenticated, IsTenantAdmin]
    
    def get_queryset(self):
        if self.request.user.role == 'super_admin':
            return TenantMailSecret.objects.all()
        elif self.request.user.tenant:
            return TenantMailSecret.objects.filter(tenant_email__tenant=self.request.user.tenant)
        return TenantMailSecret.objects.none()


# Admin ViewSet for tenant-admin (full CRUD access)
class TenantAdminGroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated, IsTenantAdmin]
    
    def get_queryset(self):
        if self.request.user.role in ['super_admin', 'support_team']:
            return Group.objects.all()
        elif self.request.user.tenant:
            return Group.objects.filter(tenant=self.request.user.tenant)
        return Group.objects.none()
    
    def perform_create(self, serializer):
        if self.request.user.role != 'super_admin':
            if not self.request.user.tenant:
                from rest_framework import serializers as rest_serializers
                raise rest_serializers.ValidationError("User must belong to a tenant")
            serializer.save(tenant=self.request.user.tenant)
        else:
            # Super admin must specify a tenant when creating groups
            tenant_id = self.request.data.get('tenant')
            if not tenant_id:
                from rest_framework import serializers as rest_serializers
                raise rest_serializers.ValidationError("Super admin must specify a tenant when creating groups")
            from .models import Tenant
            try:
                tenant = Tenant.objects.get(id=tenant_id)
                serializer.save(tenant=tenant)
            except Tenant.DoesNotExist:
                from rest_framework import serializers as rest_serializers
                raise rest_serializers.ValidationError("Invalid tenant specified")
    
    @action(detail=True, methods=['post'])
    def bulk_add_emails(self, request, pk=None):
        group = self.get_object()
        serializer = BulkEmailUploadSerializer(data=request.data)
        
        if serializer.is_valid():
            contacts = serializer.validated_data['contacts']
            
            created_count = 0
            duplicate_count = 0
            
            with transaction.atomic():
                for contact in contacts:
                    email = contact['email']
                    name = contact.get('name', '')
                    organization = contact.get('organization', '')
                    
                    group_email, created = GroupEmail.objects.get_or_create(
                        group=group,
                        email=email,
                        defaults={
                            'name': name,
                            'company': organization
                        }
                    )
                    
                    if created:
                        created_count += 1
                        
                        # Also create recipient in Recipients tab
                        from core.models_recipients import Recipient
                        recipient, recipient_created = Recipient.objects.get_or_create(
                            email=email,
                            tenant=group.tenant,
                            defaults={
                                'name': name,
                                'organization_name': contact.get('organization', ''),
                                'created_by': request.user
                            }
                        )
                        
                        # Add recipient to the contact group if it was created
                        if recipient_created:
                            from core.models_recipients import ContactGroup
                            contact_group, _ = ContactGroup.objects.get_or_create(
                                name=group.name,
                                tenant=group.tenant,
                                defaults={'created_by': request.user}
                            )
                            recipient.groups.add(contact_group)
                    else:
                        duplicate_count += 1
            
            return Response({
                'message': f'Added {created_count} emails successfully',
                'created': created_count,
                'duplicates': duplicate_count
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def upload_excel(self, request, pk=None):
        group = self.get_object()
        serializer = ExcelUploadSerializer(data=request.data)
        
        if serializer.is_valid():
            contacts = serializer.validated_data['file']
            
            created_count = 0
            duplicate_count = 0
            
            with transaction.atomic():
                for contact in contacts:
                    email = contact['email']
                    name = contact.get('name', '')
                    organization = contact.get('organization', '')
                    
                    group_email, created = GroupEmail.objects.get_or_create(
                        group=group,
                        email=email,
                        defaults={
                            'name': name,
                            'company': organization
                        }
                    )
                    
                    if created:
                        created_count += 1
                        
                        # Also create recipient in Recipients tab
                        from core.models_recipients import Recipient
                        recipient, recipient_created = Recipient.objects.get_or_create(
                            email=email,
                            tenant=group.tenant,
                            defaults={
                                'name': name,
                                'organization_name': organization,
                                'created_by': request.user
                            }
                        )
                        
                        # Add recipient to the contact group if it was created
                        if recipient_created:
                            from core.models_recipients import ContactGroup
                            contact_group, _ = ContactGroup.objects.get_or_create(
                                name=group.name,
                                tenant=group.tenant,
                                defaults={'created_by': request.user}
                            )
                            recipient.groups.add(contact_group)
                    else:
                        duplicate_count += 1
            
            return Response({
                'message': f'Added {created_count} emails successfully from Excel file',
                'created': created_count,
                'duplicates': duplicate_count
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['delete'])
    def bulk_remove_emails(self, request, pk=None):
        group = self.get_object()
        emails = request.data.get('emails', [])
        
        if not emails:
            return Response({'error': 'No emails provided'}, status=status.HTTP_400_BAD_REQUEST)
        
        deleted_count = GroupEmail.objects.filter(
            group=group,
            email__in=emails
        ).delete()[0]
        
        return Response({
            'message': f'Removed {deleted_count} emails successfully',
            'deleted': deleted_count
        })
    
    @action(detail=False, methods=['get'])
    def download_excel_template(self, request):
        """Download Excel template for bulk email upload"""
        # Create a new workbook and select the active sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Email Template"
        
        # Define headers
        headers = ['Email', 'Name', 'Organization']
        
        # Set up header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add instruction row for users
        instruction_data = [
            'Enter email addresses here', 'Enter contact names here', 'Enter company/organization names here'
        ]
        
        for col_num, instruction in enumerate(instruction_data, 1):
            cell = worksheet.cell(row=2, column=col_num, value=instruction)
            cell.font = Font(italic=True, color="666666")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response with Excel file
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="email_template.xlsx"'
        
        return response

# Staff ViewSet for tenant-staff (read-only access)
class TenantStaffGroupViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]
    
    def get_queryset(self):
        if self.request.user.role in ['super_admin', 'support_team']:
            return Group.objects.all()
        elif self.request.user.tenant:
            return Group.objects.filter(tenant=self.request.user.tenant)
        return Group.objects.none()

# Legacy GroupViewSet for backward compatibility (will be deprecated)
class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.role in ['super_admin', 'support_team']:
            return Group.objects.all()
        elif self.request.user.tenant:
            return Group.objects.filter(tenant=self.request.user.tenant)
        return Group.objects.none()
    
    def perform_create(self, serializer):
        if self.request.user.role != 'super_admin':
            if not self.request.user.tenant:
                from rest_framework import serializers as rest_serializers
                raise rest_serializers.ValidationError("User must belong to a tenant")
            serializer.save(tenant=self.request.user.tenant)
        else:
            serializer.save()


# Admin GroupEmail ViewSet for tenant-admin (full CRUD access)
class TenantAdminGroupEmailViewSet(viewsets.ModelViewSet):
    serializer_class = GroupEmailModelSerializer
    permission_classes = [IsAuthenticated, IsTenantAdmin]
    
    def get_queryset(self):
        # First apply tenant filtering
        if self.request.user.role in ['super_admin', 'support_team']:
            queryset = GroupEmail.objects.all()
        elif self.request.user.tenant:
            queryset = GroupEmail.objects.filter(group__tenant=self.request.user.tenant)
        else:
            queryset = GroupEmail.objects.none()
        
        # Then apply group filtering if group parameter is provided
        group_id = self.request.query_params.get('group', None)
        if group_id:
            queryset = queryset.filter(group_id=group_id)
        
        return queryset
    
    def perform_create(self, serializer):
        """Create group email and also add to Recipients tab"""
        group_email = serializer.save()
        
        # Also create recipient in Recipients tab
        from core.models_recipients import Recipient
        recipient, recipient_created = Recipient.objects.get_or_create(
            email=group_email.email,
            tenant=group_email.group.tenant,
            defaults={
                'name': group_email.name or '',
                'organization_name': group_email.company or '',
                'created_by': self.request.user
            }
        )
        
        # Add recipient to the contact group if it was created
        if recipient_created:
            from core.models_recipients import ContactGroup
            contact_group, _ = ContactGroup.objects.get_or_create(
                name=group_email.group.name,
                tenant=group_email.group.tenant,
                defaults={'created_by': self.request.user}
            )
            recipient.groups.add(contact_group)


# Staff GroupEmail ViewSet for tenant-staff (read-only access)
class TenantStaffGroupEmailViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = GroupEmailModelSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]
    
    def get_queryset(self):
        # First apply tenant filtering
        if self.request.user.role in ['super_admin', 'support_team']:
            queryset = GroupEmail.objects.all()
        elif self.request.user.tenant:
            queryset = GroupEmail.objects.filter(group__tenant=self.request.user.tenant)
        else:
            queryset = GroupEmail.objects.none()
        
        # Then apply group filtering if group parameter is provided
        group_id = self.request.query_params.get('group', None)
        if group_id:
            queryset = queryset.filter(group_id=group_id)
        
        return queryset


# Legacy GroupEmail ViewSet for backward compatibility
class GroupEmailViewSet(viewsets.ModelViewSet):
    serializer_class = GroupEmailModelSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]
    
    def get_queryset(self):
        # First apply tenant filtering
        if self.request.user.role in ['super_admin', 'support_team']:
            queryset = GroupEmail.objects.all()
        elif self.request.user.tenant:
            queryset = GroupEmail.objects.filter(group__tenant=self.request.user.tenant)
        else:
            queryset = GroupEmail.objects.none()
        
        # Then apply group filtering if group parameter is provided
        group_id = self.request.query_params.get('group', None)
        if group_id:
            queryset = queryset.filter(group_id=group_id)
        
        return queryset
    
    def perform_create(self, serializer):
        """Create group email and also add to Recipients tab"""
        group_email = serializer.save()
        
        # Also create recipient in Recipients tab
        from core.models_recipients import Recipient
        recipient, recipient_created = Recipient.objects.get_or_create(
            email=group_email.email,
            tenant=group_email.group.tenant,
            defaults={
                'name': group_email.name or '',
                'organization_name': group_email.company or '',
                'created_by': self.request.user
            }
        )
        
        # Add recipient to the contact group if it was created
        if recipient_created:
            from core.models_recipients import ContactGroup
            contact_group, _ = ContactGroup.objects.get_or_create(
                name=group_email.group.name,
                tenant=group_email.group.tenant,
                defaults={'created_by': self.request.user}
            )
            recipient.groups.add(contact_group)


# Role-based Batch ViewSets
class TenantAdminBatchViewSet(viewsets.ModelViewSet):
    """Batch ViewSet for tenant admin and sales team (full access)"""
    serializer_class = BatchSerializer
    permission_classes = [IsAuthenticated, IsTenantAdmin]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return BatchCreateSerializer
        return BatchSerializer
    
    def get_queryset(self):
        queryset = Batch.objects.select_related('tenant', 'template').prefetch_related('batch_groups__group')
        
        if self.request.user.role in ['super_admin', 'support_team']:
            return queryset
        elif self.request.user.tenant:
            queryset = queryset.filter(tenant=self.request.user.tenant)
        else:
            return Batch.objects.none()
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        search = self.request.query_params.get('search')
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(description__icontains=search)
            )
        
        return queryset.order_by('-created_at')
    
    def perform_create(self, serializer):
        serializer.save(tenant=self.request.user.tenant)

    @action(detail=True, methods=['post'])
    def execute_action(self, request, pk=None):
        """Execute batch actions like start, pause, resume, cancel"""
        batch = self.get_object()
        serializer = BatchActionSerializer(data=request.data, context={'batch': batch})
        
        if serializer.is_valid():
            from django.utils import timezone
            from batches.tasks import send_batch_emails, execute_batch_subcycle
            
            action_type = serializer.validated_data['action']
            
            if action_type == 'start':
                batch.status = 'scheduled'
                if batch.start_time <= timezone.now():
                    if batch.sub_cycle_enabled:
                        # Use new sub-cycle system
                        execute_batch_subcycle.delay(batch.id)
                    else:
                        # Use legacy system
                        send_batch_emails.delay(batch.id)
            
            elif action_type == 'pause':
                batch.status = 'paused'
            
            elif action_type == 'resume':
                batch.status = 'scheduled'
                if batch.start_time <= timezone.now():
                    if batch.sub_cycle_enabled:
                        # Use new sub-cycle system
                        execute_batch_subcycle.delay(batch.id)
                    else:
                        # Use legacy system
                        send_batch_emails.delay(batch.id)
            
            elif action_type == 'cancel':
                batch.status = 'cancelled'
            
            batch.save()
            
            return Response({
                'message': f'Batch {action_type}ed successfully',
                'status': batch.status
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def recipients(self, request, pk=None):
        """Get all recipients for a batch with their document status"""
        from batches.serializers import BatchRecipientSerializer
        
        batch = self.get_object()
        queryset = batch.batch_recipients.all()
        
        # Filtering
        documents_received = request.query_params.get('documents_received')
        if documents_received is not None:
            queryset = queryset.filter(documents_received=documents_received.lower() == 'true')
        
        email_sent = request.query_params.get('email_sent')
        if email_sent is not None:
            queryset = queryset.filter(email_sent=email_sent.lower() == 'true')
        
        # Search
        search = request.query_params.get('search')
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(recipient__name__icontains=search) |
                Q(recipient__email__icontains=search) |
                Q(recipient__organization_name__icontains=search)
            )
        
        page = self.paginate_queryset(queryset)
        serializer = BatchRecipientSerializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Get batch statistics"""
        batch = self.get_object()
        
        total_recipients = batch.total_recipients
        emails_sent = batch.emails_sent
        emails_failed = batch.emails_failed
        pending = total_recipients - emails_sent - emails_failed
        
        stats = {
            'total_recipients': total_recipients,
            'emails_sent': emails_sent,
            'emails_failed': emails_failed,
            'emails_pending': pending,
            'success_rate': (emails_sent / total_recipients * 100) if total_recipients > 0 else 0,
            'completion_rate': ((emails_sent + emails_failed) / total_recipients * 100) if total_recipients > 0 else 0
        }
        
        return Response(stats)

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Get dashboard statistics for batches"""
        queryset = self.get_queryset()
        
        stats = {
            'total_batches': queryset.count(),
            'active_batches': queryset.filter(status__in=['scheduled', 'running']).count(),
            'completed_batches': queryset.filter(status='completed').count(),
            'failed_batches': queryset.filter(status='failed').count(),
            'total_emails_sent': sum(batch.emails_sent for batch in queryset),
            'total_emails_failed': sum(batch.emails_failed for batch in queryset),
        }
        
        recent_batches = queryset[:5]
        stats['recent_batches'] = BatchSerializer(recent_batches, many=True).data
        
        return Response(stats)

    @action(detail=True, methods=['post'])
    def mark_documents_received(self, request, pk=None):
        """Mark documents as received for specific recipients"""
        batch = self.get_object()
        recipient_ids = request.data.get('recipient_ids', [])
        
        if not recipient_ids:
            return Response({'error': 'recipient_ids required'}, status=status.HTTP_400_BAD_REQUEST)
        
        updated = batch.batch_recipients.filter(
            recipient__in=recipient_ids
        ).update(
            is_completed=True,
            completed_at=timezone.now(),
            documents_received=True
        )
        
        # Check if batch should auto-complete
        if batch.should_auto_complete():
            batch.status = 'completed'
            batch.save()
        
        return Response({
            'updated': updated,
            'batch_status': batch.status
        })

    @action(detail=True, methods=['post'])
    def mark_documents_not_received(self, request, pk=None):
        """Mark documents as not received for specific recipients"""
        batch = self.get_object()
        recipient_ids = request.data.get('recipient_ids', [])
        
        if not recipient_ids:
            return Response({'error': 'recipient_ids required'}, status=status.HTTP_400_BAD_REQUEST)
        
        updated = batch.batch_recipients.filter(
            recipient__in=recipient_ids
        ).update(
            is_completed=False,
            completed_at=None,
            documents_received=False
        )
        
        return Response({'updated': updated})

    @action(detail=True, methods=['post'])
    def update_recipients(self, request, pk=None):
        """Update recipients for this batch"""
        batch = self.get_object()
        
        recipient_ids = request.data.get('recipient_ids', [])
        contact_group_ids = request.data.get('contact_group_ids', [])
        
        # Clear existing recipients
        batch.batch_recipients.all().delete()
        
        # Add new recipients using the serializer logic
        from batches.serializers import BatchSerializer
        serializer = BatchSerializer()
        serializer._add_recipients_to_batch(batch, recipient_ids, contact_group_ids)
        
        # Update total recipients count
        batch.total_recipients = serializer._calculate_total_recipients(batch)
        batch.save()
        
        return Response({
            'message': 'Recipients updated successfully',
            'total_recipients': batch.total_recipients
        })


class TenantStaffBatchViewSet(viewsets.ReadOnlyModelViewSet):
    """Batch ViewSet for staff admin and staff (read-only access)"""
    serializer_class = BatchSerializer
    permission_classes = [IsAuthenticated, IsTenantMember]
    
    def get_queryset(self):
        queryset = Batch.objects.select_related('tenant', 'template').prefetch_related('batch_groups__group')
        
        if self.request.user.role in ['super_admin', 'support_team']:
            return queryset
        elif self.request.user.tenant:
            queryset = queryset.filter(tenant=self.request.user.tenant)
        else:
            return Batch.objects.none()
        
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        search = self.request.query_params.get('search')
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(description__icontains=search)
            )
        
        return queryset.order_by('-created_at')

    @action(detail=True, methods=['get'])
    def recipients(self, request, pk=None):
        """Get all recipients for a batch with their document status (read-only)"""
        from batches.serializers import BatchRecipientSerializer
        
        batch = self.get_object()
        queryset = batch.batch_recipients.all()
        
        # Filtering
        documents_received = request.query_params.get('documents_received')
        if documents_received is not None:
            queryset = queryset.filter(documents_received=documents_received.lower() == 'true')
        
        email_sent = request.query_params.get('email_sent')
        if email_sent is not None:
            queryset = queryset.filter(email_sent=email_sent.lower() == 'true')
        
        # Search
        search = request.query_params.get('search')
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(recipient__name__icontains=search) |
                Q(recipient__email__icontains=search) |
                Q(recipient__organization_name__icontains=search)
            )
        
        page = self.paginate_queryset(queryset)
        serializer = BatchRecipientSerializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Get batch statistics (read-only)"""
        batch = self.get_object()
        
        total_recipients = batch.total_recipients
        emails_sent = batch.emails_sent
        emails_failed = batch.emails_failed
        pending = total_recipients - emails_sent - emails_failed
        
        stats = {
            'total_recipients': total_recipients,
            'emails_sent': emails_sent,
            'emails_failed': emails_failed,
            'emails_pending': pending,
            'success_rate': (emails_sent / total_recipients * 100) if total_recipients > 0 else 0,
            'completion_rate': ((emails_sent + emails_failed) / total_recipients * 100) if total_recipients > 0 else 0
        }
        
        return Response(stats)

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Get dashboard statistics for batches (read-only)"""
        queryset = self.get_queryset()
        
        stats = {
            'total_batches': queryset.count(),
            'active_batches': queryset.filter(status__in=['scheduled', 'running']).count(),
            'completed_batches': queryset.filter(status='completed').count(),
            'failed_batches': queryset.filter(status='failed').count(),
            'total_emails_sent': sum(batch.emails_sent for batch in queryset),
            'total_emails_failed': sum(batch.emails_failed for batch in queryset),
        }
        
        recent_batches = queryset[:5]
        stats['recent_batches'] = BatchSerializer(recent_batches, many=True).data
        
        return Response(stats)

    @action(detail=True, methods=['post'])
    def mark_documents_received(self, request, pk=None):
        """Mark documents as received for specific recipients"""
        batch = self.get_object()
        recipient_ids = request.data.get('recipient_ids', [])
        
        if not recipient_ids:
            return Response({'error': 'recipient_ids required'}, status=status.HTTP_400_BAD_REQUEST)
        
        updated = batch.batch_recipients.filter(
            recipient__in=recipient_ids
        ).update(
            is_completed=True,
            completed_at=timezone.now(),
            documents_received=True
        )
        
        # Check if batch should auto-complete
        if batch.should_auto_complete():
            batch.status = 'completed'
            batch.save()
        
        return Response({
            'updated': updated,
            'batch_status': batch.status
        })

    @action(detail=True, methods=['post'])
    def mark_documents_not_received(self, request, pk=None):
        """Mark documents as not received for specific recipients"""
        batch = self.get_object()
        recipient_ids = request.data.get('recipient_ids', [])
        
        if not recipient_ids:
            return Response({'error': 'recipient_ids required'}, status=status.HTTP_400_BAD_REQUEST)
        
        updated = batch.batch_recipients.filter(
            recipient__in=recipient_ids
        ).update(
            is_completed=False,
            completed_at=None,
            documents_received=False
        )
        
        return Response({'updated': updated})


# Tenant-specific Recipients and Contact Groups ViewSets

class TenantAdminRecipientViewSet(viewsets.ModelViewSet):
    """ViewSet for managing recipients - Tenant Admin role"""
    permission_classes = [IsAuthenticated, IsTenantAdmin]
    filterset_fields = ['is_active', 'groups']
    search_fields = ['name', 'organization_name', 'email']
    ordering_fields = ['name', 'organization_name', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        from core.models_recipients import Recipient
        if self.request.user.tenant:
            return Recipient.objects.filter(tenant=self.request.user.tenant)
        return Recipient.objects.none()
    
    def get_serializer_class(self):
        from core.serializers_recipients import RecipientSerializer
        return RecipientSerializer
    
    def perform_create(self, serializer):
        serializer.save(tenant=self.request.user.tenant, created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        self.perform_bulk_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_bulk_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download an Excel template for recipient upload"""
        try:
            from core.models_recipients import Recipient
            columns = [
                'Name*', 'Email*', 'Organization Name*', 'Title', 
                'Phone Number', 'Notes', 'Groups'
            ]
            df = pd.DataFrame(columns=columns)
            
            # Add example row
            df.loc[0] = [
                'John Doe', 'john@example.com', 'ACME Corp', 
                'Manager', '+1234567890', 'VIP Client', 'Group1, Group2'
            ]
            
            excel_file = BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Recipients')
                worksheet = writer.sheets['Recipients']
                
                # Apply styling to make template more user-friendly
                from openpyxl.styles import Font, PatternFill
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add column descriptions in a new sheet
                descriptions = pd.DataFrame([
                    ['Name*', 'Required. Full name of the recipient'],
                    ['Email*', 'Required. Valid email address'],
                    ['Organization Name*', 'Required. Company or organization name'],
                    ['Title', 'Optional. Job title or position'],
                    ['Phone Number', 'Optional. Contact number with country code'],
                    ['Notes', 'Optional. Additional information'],
                    ['Groups', 'Optional. Comma-separated group names (existing groups only)']
                ], columns=['Column', 'Description'])
                descriptions.to_excel(writer, sheet_name='Instructions', index=False)
            
            excel_file.seek(0)
            
            # Use Django's HttpResponse for better file handling
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="recipient_template.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Cache-Control'] = 'no-cache'
            
            print(f"[OK] Template download requested by user: {request.user.username}")
            return response
            
        except Exception as e:
            print(f"[ERROR] Error generating template: {str(e)}")
            return Response(
                {'error': f'Failed to generate template: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], parser_classes=[parsers.MultiPartParser])
    def upload_excel(self, request):
        from core.models_recipients import Recipient, ContactGroup
        print(f"[UPLOAD] Upload request received from user: {request.user.username}")
        print(f"[UPLOAD] Request files: {list(request.FILES.keys())}")
        
        if 'file' not in request.FILES:
            print("[ERROR] No file in request")
            return Response(
                {'error': 'Please upload a file'},
                status=status.HTTP_400_BAD_REQUEST
            )

        excel_file = request.FILES['file']
        print(f"[UPLOAD] File details: {excel_file.name}, Size: {excel_file.size} bytes")
        
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            print(f"[ERROR] Invalid file type: {excel_file.name}")
            return Response(
                {'error': 'Please upload an Excel file (.xls or .xlsx)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if excel_file.size > 10 * 1024 * 1024:  # 10MB limit
            print(f"[ERROR] File too large: {excel_file.size} bytes")
            return Response(
                {'error': 'File size too large. Maximum 10MB allowed.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Read Excel file
            df = pd.read_excel(excel_file)
            required_columns = ['Name', 'Email', 'Organization Name']
            optional_columns = ['Title', 'Phone Number', 'Notes', 'Groups']
            
            # Convert column names to lowercase for case-insensitive matching
            df.columns = [col.strip().lower() for col in df.columns]
            required_columns_lower = [col.lower() for col in required_columns]
            
            # Validate required columns
            missing_columns = [col for col in required_columns_lower if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process rows in batches
            batch_size = 100
            recipients_to_create = []
            errors = []
            created_recipients = []
            
            for index, row in df.iterrows():
                try:
                    # Validate email
                    email = str(row['email']).strip()
                    validate_email(email)
                    
                    # Check if email already exists in tenant
                    if Recipient.objects.filter(email=email, tenant=request.user.tenant).exists():
                        errors.append(f'Row {index + 2}: Email {email} already exists')
                        continue
                    
                    # Prepare recipient data
                    recipient_data = {
                        'name': str(row['name']).strip(),
                        'organization_name': str(row['organization name']).strip(),
                        'email': email,
                        'tenant': request.user.tenant,
                        'created_by': request.user
                    }
                    
                    # Add optional fields if present
                    if 'title' in df.columns and pd.notna(row['title']):
                        recipient_data['title'] = str(row['title']).strip()
                    if 'phone number' in df.columns and pd.notna(row['phone number']):
                        recipient_data['phone_number'] = str(row['phone number']).strip()
                    if 'notes' in df.columns and pd.notna(row['notes']):
                        recipient_data['notes'] = str(row['notes']).strip()
                    
                    recipients_to_create.append(recipient_data)
                    
                    # Process in batches
                    if len(recipients_to_create) >= batch_size:
                        batch_recipients = Recipient.objects.bulk_create([
                            Recipient(**data) for data in recipients_to_create
                        ])
                        created_recipients.extend(batch_recipients)
                        recipients_to_create = []
                        
                except ValidationError:
                    errors.append(f'Row {index + 2}: Invalid email format - {email}')
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')

            # Create remaining recipients
            if recipients_to_create:
                batch_recipients = Recipient.objects.bulk_create([
                    Recipient(**data) for data in recipients_to_create
                ])
                created_recipients.extend(batch_recipients)

            # Process groups if present
            if 'groups' in df.columns:
                for recipient in created_recipients:
                    try:
                        row = df[df['email'] == recipient.email].iloc[0]
                        if pd.notna(row['groups']):
                            group_names = [name.strip() for name in str(row['groups']).split(',')]
                            existing_groups = ContactGroup.objects.filter(
                                tenant=request.user.tenant,
                                name__in=group_names
                            )
                            recipient.groups.add(*existing_groups)
                    except (IndexError, Exception) as e:
                        errors.append(f'Error adding groups for {recipient.email}: {str(e)}')

            # Return results with detailed summary
            result_data = {
                'success': True,
                'created_count': len(created_recipients),
                'total_rows_processed': len(df),
                'errors_count': len(errors),
                'errors': errors[:10],  # Limit errors in response to first 10
                'recipients': self.get_serializer(created_recipients[:5], many=True).data  # Limit recipients in response
            }
            
            if len(errors) > 10:
                result_data['errors_truncated'] = True
                result_data['total_errors'] = len(errors)
            
            print(f"[OK] Upload completed: {len(created_recipients)} created, {len(errors)} errors")
            return Response(result_data)

        except pd.errors.EmptyDataError:
            print("[ERROR] Empty Excel file")
            return Response(
                {'error': 'The uploaded file is empty or contains no data'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except pd.errors.ExcelFileError:
            print("[ERROR] Invalid Excel file format")
            return Response(
                {'error': 'Invalid Excel file format. Please check your file and try again.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            print(f"[ERROR] Unexpected error processing file: {str(e)}")
            return Response(
                {'error': f'Error processing file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            export_format = request.GET.get('format', 'excel').lower()
            
            if not queryset.exists():
                return Response(
                    {'error': 'No recipients found to export'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Create DataFrame
            data = []
            for recipient in queryset:
                data.append({
                    'Name': recipient.name,
                    'Email': recipient.email,
                    'Organization Name': recipient.organization_name,
                    'Title': recipient.title or '',
                    'Phone Number': recipient.phone_number or '',
                    'Notes': recipient.notes or '',
                    'Groups': ', '.join(g.name for g in recipient.groups.all()),
                    'Created At': recipient.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Created By': recipient.created_by.email if recipient.created_by else ''
                })
            
            df = pd.DataFrame(data)
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            tenant_name = request.user.tenant.name if request.user.tenant else 'system'
            
            if export_format == 'csv':
                # Export as CSV
                csv_file = BytesIO()
                df.to_csv(csv_file, index=False, encoding='utf-8')
                csv_file.seek(0)
                
                response = HttpResponse(
                    csv_file.getvalue(),
                    content_type='text/csv'
                )
                filename = f'recipients_{tenant_name}_{timestamp}.csv'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                response['Cache-Control'] = 'no-cache'
                
                print(f"[OK] Recipients CSV export requested by user: {request.user.username} - {len(data)} records")
                return response
            
            else:
                # Export as Excel (default)
                excel_file = BytesIO()
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Recipients')
                    worksheet = writer.sheets['Recipients']
                    
                    # Apply enhanced formatting
                    from openpyxl.styles import Font, PatternFill
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        # Find maximum length of content in column
                        for cell in column:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        
                        # Add some padding and set width
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add filters
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Style headers
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    
                    for col_num in range(1, len(df.columns) + 1):
                        header_cell = worksheet.cell(row=1, column=col_num)
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                
                excel_file.seek(0)
                
                response = HttpResponse(
                    excel_file.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                filename = f'recipients_{tenant_name}_{timestamp}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                response['Cache-Control'] = 'no-cache'
                
                print(f"[OK] Recipients Excel export requested by user: {request.user.username} - {len(data)} records")
                return response
            
        except Exception as e:
            print(f"[ERROR] Error exporting recipients: {str(e)}")
            return Response(
                {'error': f'Failed to export recipients: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def bulk_update_groups(self, request):
        from core.models_recipients import ContactGroup
        recipient_ids = request.data.get('recipient_ids', [])
        group_ids = request.data.get('group_ids', [])
        action = request.data.get('action', 'add')  # 'add' or 'remove'

        if not recipient_ids or not group_ids:
            return Response(
                {"error": "Both recipient_ids and group_ids are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        recipients = self.get_queryset().filter(id__in=recipient_ids)
        groups = ContactGroup.objects.filter(
            id__in=group_ids,
            tenant=self.request.user.tenant
        )

        for recipient in recipients:
            if action == 'add':
                recipient.groups.add(*groups)
            else:
                recipient.groups.remove(*groups)

        serializer = self.get_serializer(recipients, many=True)
        return Response(serializer.data)


class TenantStaffRecipientViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing recipients - Staff role (read-only)"""
    permission_classes = [IsAuthenticated, IsTenantMember]
    filterset_fields = ['is_active', 'groups']
    search_fields = ['name', 'organization_name', 'email']
    ordering_fields = ['name', 'organization_name', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        from core.models_recipients import Recipient
        if self.request.user.tenant:
            return Recipient.objects.filter(tenant=self.request.user.tenant)
        return Recipient.objects.none()
    
    def get_serializer_class(self):
        from core.serializers_recipients import RecipientSerializer
        return RecipientSerializer

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download an Excel template for recipient upload - Read-only access"""
        try:
            columns = [
                'Name*', 'Email*', 'Organization Name*', 'Title', 
                'Phone Number', 'Notes', 'Groups'
            ]
            df = pd.DataFrame(columns=columns)
            
            # Add example row
            df.loc[0] = [
                'John Doe', 'john@example.com', 'ACME Corp', 
                'Manager', '+1234567890', 'VIP Client', 'Group1, Group2'
            ]
            
            excel_file = BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Recipients')
                worksheet = writer.sheets['Recipients']
                
                # Apply styling to make template more user-friendly
                from openpyxl.styles import Font, PatternFill
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add column descriptions in a new sheet
                descriptions = pd.DataFrame([
                    ['Name*', 'Required. Full name of the recipient'],
                    ['Email*', 'Required. Valid email address'],
                    ['Organization Name*', 'Required. Company or organization name'],
                    ['Title', 'Optional. Job title or position'],
                    ['Phone Number', 'Optional. Contact number with country code'],
                    ['Notes', 'Optional. Additional information'],
                    ['Groups', 'Optional. Comma-separated group names (existing groups only)']
                ], columns=['Column', 'Description'])
                descriptions.to_excel(writer, sheet_name='Instructions', index=False)
            
            excel_file.seek(0)
            
            # Use Django's HttpResponse for better file handling
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="recipient_template.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Cache-Control'] = 'no-cache'
            
            print(f"[OK] Template download requested by staff user: {request.user.username}")
            return response
            
        except Exception as e:
            print(f"[ERROR] Error generating template: {str(e)}")
            return Response(
                {'error': f'Failed to generate template: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export recipients to Excel or CSV - Staff can export for viewing"""
        try:
            queryset = self.filter_queryset(self.get_queryset())
            export_format = request.GET.get('format', 'excel').lower()
            
            if not queryset.exists():
                return Response(
                    {'error': 'No recipients found to export'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Create DataFrame
            data = []
            for recipient in queryset:
                data.append({
                    'Name': recipient.name,
                    'Email': recipient.email,
                    'Organization Name': recipient.organization_name,
                    'Title': recipient.title or '',
                    'Phone Number': recipient.phone_number or '',
                    'Notes': recipient.notes or '',
                    'Groups': ', '.join(g.name for g in recipient.groups.all()),
                    'Created At': recipient.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Created By': recipient.created_by.email if recipient.created_by else ''
                })
            
            df = pd.DataFrame(data)
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            tenant_name = request.user.tenant.name if request.user.tenant else 'system'
            
            if export_format == 'csv':
                # Export as CSV
                csv_file = BytesIO()
                df.to_csv(csv_file, index=False, encoding='utf-8')
                csv_file.seek(0)
                
                response = HttpResponse(
                    csv_file.getvalue(),
                    content_type='text/csv'
                )
                filename = f'recipients_{tenant_name}_{timestamp}.csv'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                response['Cache-Control'] = 'no-cache'
                
                print(f"[OK] Recipients CSV export requested by staff user: {request.user.username} - {len(data)} records")
                return response
            
            else:
                # Export as Excel (default)
                excel_file = BytesIO()
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Recipients')
                    worksheet = writer.sheets['Recipients']
                    
                    # Apply enhanced formatting
                    from openpyxl.styles import Font, PatternFill
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        # Find maximum length of content in column
                        for cell in column:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        
                        # Add some padding and set width
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add filters
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Style headers
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    
                    for col_num in range(1, len(df.columns) + 1):
                        header_cell = worksheet.cell(row=1, column=col_num)
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                
                excel_file.seek(0)
                
                response = HttpResponse(
                    excel_file.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                filename = f'recipients_{tenant_name}_{timestamp}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                response['Cache-Control'] = 'no-cache'
                
                print(f"[OK] Recipients Excel export requested by staff user: {request.user.username} - {len(data)} records")
                return response
            
        except Exception as e:
            print(f"[ERROR] Error exporting recipients: {str(e)}")
            return Response(
                {'error': f'Failed to export recipients: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def get_serializer_class(self):
        from core.serializers import RecipientSerializer
        return RecipientSerializer


class TenantAdminContactGroupViewSet(viewsets.ModelViewSet):
    """ViewSet for managing contact groups - Tenant Admin role"""
    permission_classes = [IsAuthenticated, IsTenantAdmin]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        from core.models_recipients import ContactGroup
        if self.request.user.tenant:
            return ContactGroup.objects.filter(tenant=self.request.user.tenant)
        return ContactGroup.objects.none()
    
    def get_serializer_class(self):
        from core.serializers_recipients import ContactGroupSerializer
        return ContactGroupSerializer
    
    def perform_create(self, serializer):
        serializer.save(tenant=self.request.user.tenant, created_by=self.request.user)

    @action(detail=True, methods=['get'])
    def recipients(self, request, pk=None):
        group = self.get_object()
        recipients = group.recipients.all()
        from core.serializers_recipients import RecipientSerializer
        serializer = RecipientSerializer(recipients, many=True)
        return Response(serializer.data)


class TenantStaffContactGroupViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing contact groups - Staff role (read-only)"""
    permission_classes = [IsAuthenticated, IsTenantMember]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    def get_queryset(self):
        from core.models_recipients import ContactGroup
        if self.request.user.tenant:
            return ContactGroup.objects.filter(tenant=self.request.user.tenant)
        return ContactGroup.objects.none()
    
    def get_serializer_class(self):
        from core.serializers_recipients import ContactGroupSerializer
        return ContactGroupSerializer

    @action(detail=True, methods=['get'])
    def recipients(self, request, pk=None):
        group = self.get_object()
        recipients = group.recipients.all()
        from core.serializers_recipients import RecipientSerializer
        serializer = RecipientSerializer(recipients, many=True)
        return Response(serializer.data)