from rest_framework import viewsets, permissions, status, parsers
from rest_framework.decorators import action
from rest_framework.response import Response
import pandas as pd
import openpyxl
from datetime import datetime
from django.utils import timezone
from io import BytesIO
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from .models_recipients import ContactGroup, Recipient
from .serializers_recipients import ContactGroupSerializer, RecipientSerializer
from .permissions import IsTenantUser, IsSuperAdminOrSupportTeam

class ContactGroupViewSet(viewsets.ModelViewSet):
    serializer_class = ContactGroupSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        Super admin and support team get global access, others need tenant access.
        """
        if self.request.user.is_authenticated and self.request.user.role in ['super_admin', 'support_team']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsTenantUser()]

    def get_queryset(self):
        if self.request.user.role in ['super_admin', 'support_team']:
            return ContactGroup.objects.all()
        return ContactGroup.objects.filter(tenant=self.request.user.tenant)

    @action(detail=True, methods=['get'])
    def recipients(self, request, pk=None):
        group = self.get_object()
        recipients = group.recipients.all()
        serializer = RecipientSerializer(recipients, many=True)
        return Response(serializer.data)

class RecipientViewSet(viewsets.ModelViewSet):
    serializer_class = RecipientSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['is_active', 'groups']
    search_fields = ['name', 'organization_name', 'email']
    ordering_fields = ['name', 'organization_name', 'created_at']
    ordering = ['name']

    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        Super admin and support team get global access, others need tenant access.
        """
        if self.request.user.is_authenticated and self.request.user.role in ['super_admin', 'support_team']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsTenantUser()]

    def get_queryset(self):
        if self.request.user.role in ['super_admin', 'support_team']:
            return Recipient.objects.all()
        return Recipient.objects.filter(tenant=self.request.user.tenant)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        serializer = self.get_serializer(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        self.perform_bulk_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def perform_bulk_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download an Excel template for recipient upload"""
        try:
            columns = [
                'Name*', 'Email*', 'Organization Name*', 'Title', 
                'Phone Number', 'Notes', 'Groups'
            ]
            df = pd.DataFrame(columns=columns)
            
            # Add example row
            df.loc[0] = [
                'John Doe', 'john@example.com', 'ACME Corp', 
                'Manager', '+1234567890', 'VIP Client', 'Group1, Group2'
            ]
            
            excel_file = BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Recipients')
                worksheet = writer.sheets['Recipients']
                
                # Apply styling to make template more user-friendly
                from openpyxl.styles import Font, PatternFill, Border, Side
                
                # Style headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add column descriptions in a new sheet
                descriptions = pd.DataFrame([
                    ['Name*', 'Required. Full name of the recipient'],
                    ['Email*', 'Required. Valid email address'],
                    ['Organization Name*', 'Required. Company or organization name'],
                    ['Title', 'Optional. Job title or position'],
                    ['Phone Number', 'Optional. Contact number with country code'],
                    ['Notes', 'Optional. Additional information'],
                    ['Groups', 'Optional. Comma-separated group names (existing groups only)']
                ], columns=['Column', 'Description'])
                descriptions.to_excel(writer, sheet_name='Instructions', index=False)
            
            excel_file.seek(0)
            
            # Use Django's HttpResponse for better file handling
            from django.http import HttpResponse
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="recipient_template.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Cache-Control'] = 'no-cache'
            
            print(f"[OK] Template download requested by user: {request.user.username}")
            return response
            
        except Exception as e:
            print(f"[ERROR] Error generating template: {str(e)}")
            return Response(
                {'error': f'Failed to generate template: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], parser_classes=[parsers.MultiPartParser])
    def upload_excel(self, request):
        print(f"[UPLOAD] Upload request received from user: {request.user.username}")
        print(f"[UPLOAD] Request files: {list(request.FILES.keys())}")
        
        if 'file' not in request.FILES:
            print("[ERROR] No file in request")
            return Response(
                {'error': 'Please upload a file'},
                status=status.HTTP_400_BAD_REQUEST
            )

        excel_file = request.FILES['file']
        print(f"[UPLOAD] File details: {excel_file.name}, Size: {excel_file.size} bytes")
        
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            print(f"[ERROR] Invalid file type: {excel_file.name}")
            return Response(
                {'error': 'Please upload an Excel file (.xls or .xlsx)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if excel_file.size > 10 * 1024 * 1024:  # 10MB limit
            print(f"[ERROR] File too large: {excel_file.size} bytes")
            return Response(
                {'error': 'File size too large. Maximum 10MB allowed.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Read Excel file
            df = pd.read_excel(excel_file)
            required_columns = ['Name', 'Email', 'Organization Name']
            optional_columns = ['Title', 'Phone Number', 'Notes', 'Groups']
            
            # Convert column names to lowercase for case-insensitive matching
            df.columns = [col.strip().lower() for col in df.columns]
            required_columns_lower = [col.lower() for col in required_columns]
            
            # Validate required columns
            missing_columns = [col for col in required_columns_lower if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Process rows in batches
            batch_size = 100
            recipients_to_create = []
            errors = []
            created_recipients = []
            
            for index, row in df.iterrows():
                try:
                    # Validate email
                    email = str(row['email']).strip()
                    validate_email(email)
                    
                    # Check if email already exists in tenant
                    if Recipient.objects.filter(email=email, tenant=request.user.tenant).exists():
                        errors.append(f'Row {index + 2}: Email {email} already exists')
                        continue
                    
                    # Prepare recipient data
                    recipient_data = {
                        'name': str(row['name']).strip(),
                        'organization_name': str(row['organization name']).strip(),
                        'email': email,
                        'tenant': request.user.tenant,
                        'created_by': request.user
                    }
                    
                    # Add optional fields if present
                    if 'title' in df.columns and pd.notna(row['title']):
                        recipient_data['title'] = str(row['title']).strip()
                    if 'phone number' in df.columns and pd.notna(row['phone number']):
                        recipient_data['phone_number'] = str(row['phone number']).strip()
                    if 'notes' in df.columns and pd.notna(row['notes']):
                        recipient_data['notes'] = str(row['notes']).strip()
                    
                    recipients_to_create.append(recipient_data)
                    
                    # Process in batches
                    if len(recipients_to_create) >= batch_size:
                        batch_recipients = Recipient.objects.bulk_create([
                            Recipient(**data) for data in recipients_to_create
                        ])
                        created_recipients.extend(batch_recipients)
                        recipients_to_create = []
                        
                except ValidationError:
                    errors.append(f'Row {index + 2}: Invalid email format - {email}')
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')

            # Create remaining recipients
            if recipients_to_create:
                batch_recipients = Recipient.objects.bulk_create([
                    Recipient(**data) for data in recipients_to_create
                ])
                created_recipients.extend(batch_recipients)

            # Process groups if present
            if 'groups' in df.columns:
                for recipient in created_recipients:
                    try:
                        row = df[df['email'] == recipient.email].iloc[0]
                        if pd.notna(row['groups']):
                            group_names = [name.strip() for name in str(row['groups']).split(',')]
                            existing_groups = ContactGroup.objects.filter(
                                tenant=request.user.tenant,
                                name__in=group_names
                            )
                            recipient.groups.add(*existing_groups)
                    except (IndexError, Exception) as e:
                        errors.append(f'Error adding groups for {recipient.email}: {str(e)}')

            # Return results with detailed summary
            result_data = {
                'success': True,
                'created_count': len(created_recipients),
                'total_rows_processed': len(df),
                'errors_count': len(errors),
                'errors': errors[:10],  # Limit errors in response to first 10
                'recipients': RecipientSerializer(created_recipients[:5], many=True).data  # Limit recipients in response
            }
            
            if len(errors) > 10:
                result_data['errors_truncated'] = True
                result_data['total_errors'] = len(errors)
            
            print(f"[OK] Upload completed: {len(created_recipients)} created, {len(errors)} errors")
            return Response(result_data)

        except pd.errors.EmptyDataError:
            print("[ERROR] Empty Excel file")
            return Response(
                {'error': 'The uploaded file is empty or contains no data'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except pd.errors.ExcelFileError:
            print("[ERROR] Invalid Excel file format")
            return Response(
                {'error': 'Invalid Excel file format. Please check your file and try again.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            print(f"[ERROR] Unexpected error processing file: {str(e)}")
            return Response(
                {'error': f'Error processing file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        try:
            queryset = self.filter_queryset(self.get_queryset())
            export_format = request.GET.get('format', 'excel').lower()
            
            if not queryset.exists():
                return Response(
                    {'error': 'No recipients found to export'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Create DataFrame
            data = []
            for recipient in queryset:
                data.append({
                    'Name': recipient.name,
                    'Email': recipient.email,
                    'Organization Name': recipient.organization_name,
                    'Title': recipient.title or '',
                    'Phone Number': recipient.phone_number or '',
                    'Notes': recipient.notes or '',
                    'Groups': ', '.join(g.name for g in recipient.groups.all()),
                    'Created At': recipient.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    'Created By': recipient.created_by.email if recipient.created_by else ''
                })
            
            df = pd.DataFrame(data)
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            tenant_name = request.user.tenant.name if request.user.tenant else 'system'
            
            from django.http import HttpResponse
            
            if export_format == 'csv':
                # Export as CSV
                csv_file = BytesIO()
                df.to_csv(csv_file, index=False, encoding='utf-8')
                csv_file.seek(0)
                
                response = HttpResponse(
                    csv_file.getvalue(),
                    content_type='text/csv'
                )
                filename = f'recipients_{tenant_name}_{timestamp}.csv'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                response['Cache-Control'] = 'no-cache'
                
                print(f"[OK] Recipients CSV export requested by user: {request.user.username} - {len(data)} records")
                return response
            
            else:
                # Export as Excel (default)
                excel_file = BytesIO()
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Recipients')
                    worksheet = writer.sheets['Recipients']
                    
                    # Apply enhanced formatting
                    from openpyxl.styles import Font, PatternFill
                    
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        # Find maximum length of content in column
                        for cell in column:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        
                        # Add some padding and set width
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Add filters
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Style headers
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    
                    for col_num in range(1, len(df.columns) + 1):
                        header_cell = worksheet.cell(row=1, column=col_num)
                        header_cell.font = header_font
                        header_cell.fill = header_fill
                
                excel_file.seek(0)
                
                response = HttpResponse(
                    excel_file.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                filename = f'recipients_{tenant_name}_{timestamp}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                response['Cache-Control'] = 'no-cache'
                
                print(f"[OK] Recipients Excel export requested by user: {request.user.username} - {len(data)} records")
                return response
            
        except Exception as e:
            print(f"[ERROR] Error exporting recipients: {str(e)}")
            return Response(
                {'error': f'Failed to export recipients: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def bulk_update_groups(self, request):
        recipient_ids = request.data.get('recipient_ids', [])
        group_ids = request.data.get('group_ids', [])
        action = request.data.get('action', 'add')  # 'add' or 'remove'

        if not recipient_ids or not group_ids:
            return Response(
                {"error": "Both recipient_ids and group_ids are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        recipients = self.get_queryset().filter(id__in=recipient_ids)
        groups = ContactGroup.objects.filter(
            id__in=group_ids,
            tenant=self.request.user.tenant
        )

        for recipient in recipients:
            if action == 'add':
                recipient.groups.add(*groups)
            else:
                recipient.groups.remove(*groups)

        serializer = self.get_serializer(recipients, many=True)
        return Response(serializer.data)
